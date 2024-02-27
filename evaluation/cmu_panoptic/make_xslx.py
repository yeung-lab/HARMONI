import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define the columns and their headings
columns = [
    ("Image", 60),
    ("Parent-child touching", 20),
    ("Touching certainty", 20),
    ("Parent-child Visibility", 20),
    ("Visibility certainty", 20),
    ("Notes", 30),
    ("Filename", 20)
]

# Set column headers
for col_num, (header, width) in enumerate(columns, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header
    sheet.column_dimensions[get_column_letter(col_num)].width = width

# Set possible values for "Parent-child touching" and "Touching certainty"
parent_child_touching_values = ["Touching", "Not touching", "N/A"]
certainty_values = ["I can tell", "I cannot tell"]
parent_child_visibility_values = [
    "adult and infant can see each other", 
    "infant can see the adult",
    "adult can see the infant",
    "can't see each other",
    "N/A"]

# Path to the folder containing JPG images
image_folder = "sampled_frames_v2"

touching_validation = DataValidation(
    type="list",
    formula1=f'"{",".join(parent_child_touching_values)}"',
    showDropDown=True
)
sheet.add_data_validation(touching_validation)

visibility_validation = DataValidation(
    type="list",
    formula1=f'"{",".join(parent_child_visibility_values)}"',
    showDropDown=True
)
sheet.add_data_validation(visibility_validation)

# Add a drop-down list to "Touching/Visibility certainty" column
certainty_validation = DataValidation(
    type="list",
    formula1=f'"{",".join(certainty_values)}"',
    showDropDown=True
)
sheet.add_data_validation(certainty_validation)

# Iterate through JPG files in the folder
for idx, filename in enumerate(os.listdir(image_folder)):
    if filename.endswith(".jpg"):
        image_path = os.path.join(image_folder, filename)

        # Load the image
        img = Image.open(image_path)
        h, w = img.size
        # img = img.resize((int(h/4), int(w/4)))

        # Create an ExcelImage object from the image
        excel_image = ExcelImage(img)
        # resize the image
        excel_image.width = 300
        excel_image.height = 200

        row_num = idx + 2
        # Add the image to the worksheet
        sheet.add_image(excel_image, f'A{row_num}')
        # resize the cell
        sheet.row_dimensions[row_num].height = 200

        # Set values for other columns
        sheet.cell(row=row_num, column=2, value="")
        touching_validation.add(sheet[f"B{row_num}"])

        sheet.cell(row=row_num, column=3, value="")
        certainty_validation.add(sheet[f"C{row_num}"])

        sheet.cell(row=row_num, column=4, value="")
        visibility_validation.add(sheet[f"D{row_num}"])

        sheet.cell(row=row_num, column=5, value="")
        certainty_validation.add(sheet[f"E{row_num}"])
        
        sheet.cell(row=row_num, column=7, value=filename)

# Save the workbook as an XLSX file
workbook.save("annotation_v2.xlsx")

